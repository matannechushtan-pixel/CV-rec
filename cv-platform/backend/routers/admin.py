import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.auth import require_admin
from core.database import get_db
from models.user import User, Profile
from models.cv import CV
from models.application import Application
from models.company import Company, CompanyJobPost
from models.job import JobListing

router = APIRouter()

HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class AdminStats(BaseModel):
    total_users: int
    total_cvs: int
    total_applications: int
    total_jobs_posted: int


def _style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


@router.get("/stats", response_model=AdminStats)
async def get_stats(user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    total_users = (await db.execute(select(func.count(User.id)))).scalar() or 0
    total_cvs = (await db.execute(select(func.count(CV.id)))).scalar() or 0
    total_applications = (await db.execute(select(func.count(Application.id)))).scalar() or 0
    total_jobs_posted = (await db.execute(select(func.count(CompanyJobPost.id)))).scalar() or 0
    return AdminStats(
        total_users=total_users,
        total_cvs=total_cvs,
        total_applications=total_applications,
        total_jobs_posted=total_jobs_posted,
    )


@router.get("/users/export")
async def export_users(user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    job_seekers = (
        await db.execute(
            select(
                User.id,
                Profile.full_name,
                User.email,
                User.created_at,
                func.count(func.distinct(CV.id)).label("cv_count"),
                func.count(func.distinct(Application.id)).label("applications_count"),
            )
            .outerjoin(Profile, Profile.user_id == User.id)
            .outerjoin(CV, CV.user_id == User.id)
            .outerjoin(Application, Application.user_id == User.id)
            .where(User.role == "job_seeker")
            .group_by(User.id, Profile.full_name, User.email, User.created_at)
        )
    ).all()

    employers = (
        await db.execute(
            select(
                User.id,
                User.email,
                Company.name,
                User.created_at,
                func.count(func.distinct(CompanyJobPost.id)).label("jobs_posted"),
            )
            .outerjoin(Company, Company.admin_user_id == User.id)
            .outerjoin(CompanyJobPost, CompanyJobPost.company_id == Company.id)
            .where(User.role == "company_admin")
            .group_by(User.id, User.email, Company.name, User.created_at)
        )
    ).all()

    wb = Workbook()

    ws_seekers = wb.active
    ws_seekers.title = "Job Seekers"
    ws_seekers.append(
        ["id", "full_name", "email", "created_at", "cv_count", "applications_count", "last_login"]
    )
    for row in job_seekers:
        ws_seekers.append(
            [
                str(row.id),
                row.full_name,
                row.email,
                row.created_at.replace(tzinfo=None) if row.created_at else None,
                row.cv_count,
                row.applications_count,
                None,
            ]
        )
    _style_header(ws_seekers)

    ws_employers = wb.create_sheet("Employers")
    ws_employers.append(["id", "email", "company_name", "created_at", "jobs_posted"])
    for row in employers:
        ws_employers.append(
            [
                str(row.id),
                row.email,
                row.name,
                row.created_at.replace(tzinfo=None) if row.created_at else None,
                row.jobs_posted,
            ]
        )
    _style_header(ws_employers)

    for ws in (ws_seekers, ws_employers):
        for column_cells in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"users_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
